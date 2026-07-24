"""Excel-backed A4 shipping report inspection and PDF generation."""

from __future__ import annotations

import base64
import gc
import math
import json
import os
import re
import sys
import tempfile
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook
from PySide6.QtCore import QBuffer, QByteArray, QIODevice, QMarginsF, QPointF, QRectF, QSize, Qt
from PySide6.QtGui import (
    QColor,
    QFont,
    QFontDatabase,
    QImage,
    QPageLayout,
    QPageSize,
    QPainter,
    QPainterPath,
    QPdfWriter,
    QPen,
    QPolygonF,
)
from PySide6.QtPdf import QPdfDocument
from PySide6.QtWidgets import QApplication

from .excel_export import EXPORT_WORKBOOK_SCHEMA_VERSION


REPORT_WORKBOOK_SCHEMA_VERSION = EXPORT_WORKBOOK_SCHEMA_VERSION
REPORT_DATA_SHEET = "测试数据"
REPORT_SESSION_SHEET = "会话信息"
_REPORT_QT_APPLICATION: QApplication | None = None


class ShippingReportType(str, Enum):
    SPECTRUM = "spectrum"
    POLE = "pole"


class SpectrumAxisMode(str, Enum):
    COUNTS = "counts"
    RELATIVE_DB = "relative_db"


@dataclass(frozen=True)
class ReportFieldDefinition:
    key: str
    label: str
    unit: str
    side: str
    measured_key: str | None = None


SPECTRUM_FIELD_DEFINITIONS = (
    ReportFieldDefinition("power", "Output Power (CW)", "W", "left", "powerW"),
    ReportFieldDefinition("operatingCurrent", "Operating Current", "A", "left", "currentA"),
    ReportFieldDefinition("centerWavelength", "Center Wavelength", "nm", "left", "peakWavelengthNm"),
    ReportFieldDefinition("fwhm", "Spectral Width (FWHM)", "nm", "left", "fwhmNm"),
    ReportFieldDefinition("wavelengthCurrentCoefficient", "Wavelength-Current Coefficient", "nm/A", "left"),
    ReportFieldDefinition("coolingTemperature", "CoolingTemperature", "°C", "left"),
    ReportFieldDefinition("powerFluctuation", "Power Fluctuation", "%", "left"),
    ReportFieldDefinition("thresholdCurrent", "Threshold Current", "A", "right"),
    ReportFieldDefinition("operatingVoltage", "Operating Voltage", "V", "right", "voltageV"),
    ReportFieldDefinition("slopeEfficiency", "Slope Efficiency", "W/A", "right", "slopeEfficiencyWPerA"),
    ReportFieldDefinition("electroOpticalEfficiency", "E-O efficiency", "%", "right", "efficiencyPercent"),
    ReportFieldDefinition("na017Energy", "NA 0.17 Encircled Energy", "%", "right"),
    ReportFieldDefinition("fiberCoreDiameter", "Fiber Core Diameter", "μm", "right"),
    ReportFieldDefinition("fiberNa", "Fiber NA", "", "right"),
)

POLE_FIELD_DEFINITIONS = (
    ReportFieldDefinition("power", "Output Power", "W", "single", "powerW"),
    ReportFieldDefinition("operatingCurrent", "Operating Current", "A", "single", "currentA"),
    ReportFieldDefinition("operatingVoltage", "Operating Voltage", "V", "single", "voltageV"),
    ReportFieldDefinition("slopeEfficiency", "Slope Efficiency", "W/A", "single", "slopeEfficiencyWPerA"),
    ReportFieldDefinition("electroOpticalEfficiency", "E-O efficiency", "%", "single", "efficiencyPercent"),
    ReportFieldDefinition("coolingTemperature", "CoolingTemperature", "°C", "single"),
)

# The approved Pole report sample presents this complete six-row summary.
# Keep it as the first-use default while preserving any explicit user choice
# (including an intentionally saved empty selection) in QSettings.
DEFAULT_POLE_SELECTED_FIELD_KEYS = tuple(field.key for field in POLE_FIELD_DEFINITIONS)


@dataclass(frozen=True)
class ReportPoint:
    current_a: float
    voltage_v: float
    power_w: float
    efficiency: float
    peak_wavelength_nm: float | None
    centroid_nm: float | None
    fwhm_nm: float | None
    pib: float | None
    smsr_db: float | None
    wavelength: tuple[float, ...]
    intensity: tuple[float, ...]

    @property
    def has_complete_spectrum(self) -> bool:
        return (
            len(self.wavelength) >= 2
            and len(self.wavelength) == len(self.intensity)
            and all(math.isfinite(value) for value in self.wavelength)
            and all(math.isfinite(value) for value in self.intensity)
        )

    def suggestions(self) -> dict[str, float | None]:
        slope_efficiency = None
        if not math.isclose(self.current_a, 0.0, abs_tol=1e-12):
            calculated = self.power_w / self.current_a
            if math.isfinite(calculated):
                slope_efficiency = calculated
        return {
            "currentA": self.current_a,
            "voltageV": self.voltage_v,
            "powerW": self.power_w,
            "slopeEfficiencyWPerA": slope_efficiency,
            "efficiencyPercent": self.efficiency * 100.0,
            "peakWavelengthNm": self.peak_wavelength_nm,
            "centroidNm": self.centroid_nm,
            "fwhmNm": self.fwhm_nm,
            "pibPercent": None if self.pib is None else self.pib * 100.0,
            "smsrDb": self.smsr_db,
        }


@dataclass(frozen=True)
class WorkbookCompatibility:
    kind: str
    status: str
    message: str
    requires_legacy_confirmation: bool


@dataclass(frozen=True)
class WorkbookInspection:
    source_path: Path
    schema_version: str
    sn: str
    product_name: str
    session: Mapping[str, str]
    points: tuple[ReportPoint, ...]
    compatibility: WorkbookCompatibility

    @property
    def spectrum_complete(self) -> bool:
        return bool(self.points) and all(point.has_complete_spectrum for point in self.points)

    @property
    def allowed_report_types(self) -> tuple[str, ...]:
        return (
            (ShippingReportType.SPECTRUM.value, ShippingReportType.POLE.value)
            if self.spectrum_complete
            else (ShippingReportType.POLE.value,)
        )

    def point_for_current(self, current_a: float) -> ReportPoint:
        requested = float(current_a)
        for point in self.points:
            if math.isclose(point.current_a, requested, rel_tol=0.0, abs_tol=1e-9):
                return point
        raise ValueError(f"Excel 中不存在 {requested:g} A 测试点")

    def to_payload(self) -> dict[str, Any]:
        return {
            "sourcePath": str(self.source_path),
            "schemaVersion": self.schema_version,
            "sn": self.sn,
            "productName": self.product_name,
            "currents": [point.current_a for point in self.points],
            "points": [
                {
                    **point.suggestions(),
                    "hasSpectrum": point.has_complete_spectrum,
                }
                for point in self.points
            ],
            "spectrumComplete": self.spectrum_complete,
            "allowedReportTypes": list(self.allowed_report_types),
            "compatibility": {
                "kind": self.compatibility.kind,
                "status": self.compatibility.status,
                "message": self.compatibility.message,
                "requiresLegacyConfirmation": self.compatibility.requires_legacy_confirmation,
            },
            "fieldDefinitions": {
                ShippingReportType.SPECTRUM.value: [_field_payload(item) for item in SPECTRUM_FIELD_DEFINITIONS],
                ShippingReportType.POLE.value: [_field_payload(item) for item in POLE_FIELD_DEFINITIONS],
            },
        }


@dataclass(frozen=True)
class SelectedReportField:
    include: bool
    value: str


@dataclass(frozen=True)
class SpectrumAxisSettings:
    mode: SpectrumAxisMode
    minimum: float
    maximum: float


@dataclass(frozen=True)
class ShippingReportRequest:
    report_type: ShippingReportType
    product_name: str
    sn: str
    operating_current_a: float
    fields: Mapping[str, SelectedReportField]
    legacy_completion_confirmed: bool
    custom_fields: tuple[ReportFieldDefinition, ...] = ()
    spectrum_axis: SpectrumAxisSettings | None = None

    @staticmethod
    def from_mapping(values: Mapping[str, Any]) -> "ShippingReportRequest":
        try:
            report_type = ShippingReportType(str(values.get("reportType", "")))
        except ValueError as exc:
            raise ValueError("请选择有效的报告类型") from exc
        raw_fields = values.get("fields")
        if not isinstance(raw_fields, Mapping):
            raise ValueError("报告参数格式无效")
        raw_custom_fields = values.get("customFields", [])
        if not isinstance(raw_custom_fields, list):
            raise ValueError("自定义报告参数格式无效")
        custom_fields: list[ReportFieldDefinition] = []
        default_side = "left" if report_type is ShippingReportType.SPECTRUM else "single"
        for raw in raw_custom_fields:
            if not isinstance(raw, Mapping):
                raise ValueError("自定义报告参数格式无效")
            custom_fields.append(
                ReportFieldDefinition(
                    key=str(raw.get("key", "")).strip(),
                    label=str(raw.get("label", "")).strip(),
                    unit=str(raw.get("unit", "")).strip(),
                    side=default_side,
                )
            )
        fields: dict[str, SelectedReportField] = {}
        for key, raw in raw_fields.items():
            if not isinstance(raw, Mapping):
                raise ValueError(f"参数 {key} 的格式无效")
            fields[str(key)] = SelectedReportField(
                include=bool(raw.get("include")),
                value=str(raw.get("value", "")).strip(),
            )
        spectrum_axis: SpectrumAxisSettings | None = None
        if report_type is ShippingReportType.SPECTRUM:
            raw_axis = values.get("spectrumAxis")
            if not isinstance(raw_axis, Mapping):
                raise ValueError("请填写光谱纵轴设置")
            try:
                spectrum_axis = SpectrumAxisSettings(
                    mode=SpectrumAxisMode(str(raw_axis.get("mode", ""))),
                    minimum=float(raw_axis.get("minimum")),
                    maximum=float(raw_axis.get("maximum")),
                )
            except (TypeError, ValueError) as exc:
                raise ValueError("光谱纵轴设置无效") from exc
        try:
            operating_current_a = float(values.get("operatingCurrentA"))
        except (TypeError, ValueError) as exc:
            raise ValueError("请选择工作电流点") from exc
        return ShippingReportRequest(
            report_type=report_type,
            product_name=str(values.get("productName", "")).strip(),
            sn=str(values.get("sn", "")).strip(),
            operating_current_a=operating_current_a,
            fields=fields,
            legacy_completion_confirmed=bool(values.get("legacyCompletionConfirmed")),
            custom_fields=tuple(custom_fields),
            spectrum_axis=spectrum_axis,
        )


def _field_payload(field: ReportFieldDefinition) -> dict[str, str | None]:
    return {
        "key": field.key,
        "label": field.label,
        "unit": field.unit,
        "side": field.side,
        "measuredKey": field.measured_key,
    }


def _finite_or_none(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _required_number(value: Any, name: str) -> float:
    result = _finite_or_none(value)
    if result is None:
        raise ValueError(f"{name} 必须是有效数值")
    return result


def _report_definitions(
    report_type: ShippingReportType,
    custom_fields: Iterable[ReportFieldDefinition] = (),
) -> tuple[ReportFieldDefinition, ...]:
    standard = SPECTRUM_FIELD_DEFINITIONS if report_type is ShippingReportType.SPECTRUM else POLE_FIELD_DEFINITIONS
    return (*standard, *tuple(custom_fields))


def _parse_session_information(workbook: Any) -> dict[str, str]:
    if REPORT_SESSION_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[REPORT_SESSION_SHEET]
    output: dict[str, str] = {}
    for row in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
        label, value = row
        if label is None:
            continue
        output[str(label).strip()] = "" if value is None else str(value).strip()
    return output


def _compatibility(session: Mapping[str, str], schema_version: str) -> WorkbookCompatibility:
    status = str(session.get("流程状态", "")).strip().lower()
    explicit_failures = {"stopped_by_operator", "aborted_safely", "failed", "error"}
    if status in explicit_failures:
        return WorkbookCompatibility(
            kind="rejected",
            status=status,
            message="该 Excel 明确记录为中止或失败测试，不能生成出货报告。",
            requires_legacy_confirmation=False,
        )
    if schema_version and schema_version != REPORT_WORKBOOK_SCHEMA_VERSION:
        return WorkbookCompatibility(
            kind="rejected",
            status=status,
            message=f"该 Excel 的出货报告格式版本 {schema_version} 不受当前软件支持。",
            requires_legacy_confirmation=False,
        )
    if schema_version == REPORT_WORKBOOK_SCHEMA_VERSION:
        if status == "completed":
            return WorkbookCompatibility(
                kind="verified_success",
                status=status,
                message="已验证为完整成功测试。",
                requires_legacy_confirmation=False,
            )
        return WorkbookCompatibility(
            kind="rejected",
            status=status,
            message="该 Excel 尚未记录完整成功状态，不能生成出货报告。",
            requires_legacy_confirmation=False,
        )
    return WorkbookCompatibility(
        kind="legacy_needs_confirmation",
        status=status,
        message="旧版 Excel 无法可靠验证完成状态，请确认该文件来自完整成功测试。",
        requires_legacy_confirmation=True,
    )


def inspect_shipping_workbook(path: Path | str) -> WorkbookInspection:
    source = Path(path).expanduser().resolve()
    if source.suffix.lower() != ".xlsx":
        raise ValueError("请选择 .xlsx 测试文件")
    if not source.is_file():
        raise ValueError("所选 Excel 文件不存在")
    try:
        # Random access is required to pair result rows with many spectrum
        # columns. Normal mode avoids repeatedly re-streaming the worksheet for
        # every cell, which becomes prohibitively slow for real spectrometers.
        workbook = load_workbook(source, read_only=False, data_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取 Excel：{exc}") from exc
    try:
        session = _parse_session_information(workbook)
        schema_version = str(session.get("出货报告格式版本", "")).strip()
        sheet = workbook[REPORT_DATA_SHEET] if REPORT_DATA_SHEET in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        points_by_current: dict[float, ReportPoint] = {}
        row = 3
        while True:
            raw_current = sheet.cell(row=row, column=1).value
            if raw_current is None or str(raw_current).strip() == "":
                break
            current = _required_number(raw_current, f"第 {row} 行电流")
            if current in points_by_current:
                raise ValueError(f"Excel 中存在重复电流点：{current:g} A")
            point = ReportPoint(
                current_a=current,
                voltage_v=_required_number(sheet.cell(row=row, column=2).value, f"{current:g} A 电压"),
                power_w=_required_number(sheet.cell(row=row, column=3).value, f"{current:g} A 功率"),
                efficiency=_required_number(sheet.cell(row=row, column=4).value, f"{current:g} A 效率"),
                peak_wavelength_nm=_finite_or_none(sheet.cell(row=row, column=5).value),
                centroid_nm=_finite_or_none(sheet.cell(row=row, column=6).value),
                fwhm_nm=_finite_or_none(sheet.cell(row=row, column=7).value),
                pib=_finite_or_none(sheet.cell(row=row, column=8).value),
                smsr_db=_finite_or_none(sheet.cell(row=row, column=9).value),
                wavelength=(),
                intensity=(),
            )
            points_by_current[current] = point
            row += 1
        if not points_by_current:
            raise ValueError("Excel 中没有有效测试点")

        spectra: dict[float, tuple[tuple[float, ...], tuple[float, ...]]] = {}
        for column in range(10, sheet.max_column + 1, 2):
            label = sheet.cell(row=2, column=column).value
            if label is None:
                continue
            match = re.fullmatch(r"\s*([+-]?(?:\d+(?:\.\d*)?|\.\d+))\s*A\s*", str(label), re.IGNORECASE)
            if match is None:
                continue
            current = float(match.group(1))
            if current in spectra:
                raise ValueError(f"Excel 中存在重复光谱列：{current:g} A")
            wavelength: list[float] = []
            intensity: list[float] = []
            invalid_pair = False
            for spectrum_row in range(3, sheet.max_row + 1):
                x = _finite_or_none(sheet.cell(row=spectrum_row, column=column).value)
                y = _finite_or_none(sheet.cell(row=spectrum_row, column=column + 1).value)
                if x is None and y is None:
                    continue
                if x is None or y is None:
                    invalid_pair = True
                    break
                wavelength.append(x)
                intensity.append(y)
            spectra[current] = ((), ()) if invalid_pair else (tuple(wavelength), tuple(intensity))

        points = []
        for current in sorted(points_by_current):
            point = points_by_current[current]
            wavelength, intensity = spectra.get(current, ((), ()))
            points.append(
                ReportPoint(
                    current_a=point.current_a,
                    voltage_v=point.voltage_v,
                    power_w=point.power_w,
                    efficiency=point.efficiency,
                    peak_wavelength_nm=point.peak_wavelength_nm,
                    centroid_nm=point.centroid_nm,
                    fwhm_nm=point.fwhm_nm,
                    pib=point.pib,
                    smsr_db=point.smsr_db,
                    wavelength=wavelength,
                    intensity=intensity,
                )
            )
        return WorkbookInspection(
            source_path=source,
            schema_version=schema_version,
            sn=session.get("SN", "").strip(),
            product_name=session.get("产品型号", "").strip(),
            session=session,
            points=tuple(points),
            compatibility=_compatibility(session, schema_version),
        )
    finally:
        workbook.close()


def validate_shipping_report_request(inspection: WorkbookInspection, request: ShippingReportRequest) -> None:
    if inspection.compatibility.kind == "rejected":
        raise ValueError(inspection.compatibility.message)
    if inspection.compatibility.requires_legacy_confirmation and not request.legacy_completion_confirmed:
        raise ValueError("请先确认旧版 Excel 来自完整成功测试")
    if not request.product_name:
        raise ValueError("产品名称不能为空")
    if not request.sn:
        raise ValueError("SN 不能为空")
    if request.report_type.value not in inspection.allowed_report_types:
        raise ValueError("该 Excel 的光谱数据不完整，只能生成 Pole 无光谱报告")
    inspection.point_for_current(request.operating_current_a)
    definitions = _report_definitions(request.report_type, request.custom_fields)
    standard_keys = {field.key for field in _report_definitions(request.report_type)}
    custom_keys = [field.key for field in request.custom_fields]
    if any(not field.key or not field.label for field in request.custom_fields):
        raise ValueError("自定义报告参数的名称不能为空")
    if len(custom_keys) != len(set(custom_keys)) or standard_keys.intersection(custom_keys):
        raise ValueError("自定义报告参数名称不能重复或使用已有参数名称")
    if len(request.custom_fields) > 30:
        raise ValueError("自定义报告参数不能超过 30 项")
    allowed_keys = {field.key for field in definitions}
    custom_key_set = set(custom_keys)
    included = []
    for key, selected in request.fields.items():
        if key not in allowed_keys:
            raise ValueError(f"未知报告参数：{key}")
        if selected.include:
            included.append(key)
            if not selected.value:
                definition = next(item for item in definitions if item.key == key)
                raise ValueError(f"请填写已勾选参数“{definition.label}”")
            if key not in custom_key_set:
                _required_number(selected.value, next(item.label for item in definitions if item.key == key))
    if not included:
        raise ValueError("请至少勾选一个报告参数")
    if request.report_type is ShippingReportType.SPECTRUM:
        axis = request.spectrum_axis
        if axis is None or not math.isfinite(axis.minimum) or not math.isfinite(axis.maximum):
            raise ValueError("请填写有效的光谱纵轴范围")
        if axis.minimum >= axis.maximum:
            raise ValueError("光谱纵轴下限必须小于上限")
        if axis.mode is SpectrumAxisMode.COUNTS and axis.minimum < 0:
            raise ValueError("原始强度纵轴下限不能小于 0")
        if axis.mode is SpectrumAxisMode.RELATIVE_DB and axis.maximum > 0:
            raise ValueError("相对强度纵轴上限不能大于 0 dB")


def suggested_field_values(
    inspection: WorkbookInspection,
    report_type: ShippingReportType,
    operating_current_a: float,
) -> dict[str, str]:
    point = inspection.point_for_current(operating_current_a)
    suggestions = point.suggestions()
    definitions = SPECTRUM_FIELD_DEFINITIONS if report_type is ShippingReportType.SPECTRUM else POLE_FIELD_DEFINITIONS
    output: dict[str, str] = {}
    for definition in definitions:
        if not definition.measured_key:
            continue
        value = suggestions.get(definition.measured_key)
        if value is None:
            continue
        if definition.key in {"operatingCurrent"}:
            output[definition.key] = f"{value:g}"
        elif definition.key in {"power"}:
            output[definition.key] = f"{value:.1f}"
        elif definition.key in {"operatingVoltage", "electroOpticalEfficiency"}:
            output[definition.key] = f"{value:.2f}"
        elif definition.key == "slopeEfficiency":
            output[definition.key] = f"{value:.3f}"
        else:
            output[definition.key] = f"{value:.3f}".rstrip("0").rstrip(".")
    return output


def load_shipping_report_preferences(settings: Any) -> dict[str, Any]:
    """Return report preferences from QSettings without trusting stored shapes."""

    def stored_float(key: str, default: float) -> float:
        try:
            value = float(settings.value(key, default))
        except (TypeError, ValueError):
            return default
        return value if math.isfinite(value) else default

    output: dict[str, Any] = {}
    for report_type in ShippingReportType:
        prefix = f"shipping_report/{report_type.value}/"
        raw_selected = settings.value(prefix + "selected_fields", None)
        try:
            selected = json.loads(str(raw_selected)) if raw_selected is not None else []
            manual_values = json.loads(str(settings.value(prefix + "manual_values", "{}")))
        except (TypeError, ValueError, json.JSONDecodeError):
            selected, manual_values = [], {}
        if not isinstance(selected, list):
            selected = []
        if raw_selected is None and report_type is ShippingReportType.POLE:
            selected = list(DEFAULT_POLE_SELECTED_FIELD_KEYS)
        if not isinstance(manual_values, dict):
            manual_values = {}
        try:
            custom_fields = json.loads(str(settings.value(prefix + "custom_fields", "[]")))
        except (TypeError, ValueError, json.JSONDecodeError):
            custom_fields = []
        if not isinstance(custom_fields, list):
            custom_fields = []
        normalized_custom_fields = []
        for item in custom_fields:
            if not isinstance(item, dict):
                continue
            key = str(item.get("key", "")).strip()
            label = str(item.get("label", "")).strip()
            if not key or not label:
                continue
            normalized_custom_fields.append({
                "key": key,
                "label": label,
                "unit": str(item.get("unit", "")).strip(),
                "side": "left" if report_type is ShippingReportType.SPECTRUM else "single",
            })
        preference: dict[str, Any] = {
            "selectedFields": [str(value) for value in selected],
            "manualValues": {str(key): str(value) for key, value in manual_values.items()},
            "customFields": normalized_custom_fields,
        }
        if report_type is ShippingReportType.SPECTRUM:
            mode = str(settings.value(prefix + "axis_mode", SpectrumAxisMode.RELATIVE_DB.value))
            if mode not in {item.value for item in SpectrumAxisMode}:
                mode = SpectrumAxisMode.RELATIVE_DB.value
            preference["spectrumAxis"] = {
                "mode": mode,
                "minimum": stored_float(prefix + "axis_minimum", -80.0),
                "maximum": stored_float(prefix + "axis_maximum", 0.0),
            }
        output[report_type.value] = preference
    return output


def save_shipping_report_preferences(settings: Any, request: ShippingReportRequest) -> None:
    definitions = _report_definitions(request.report_type, request.custom_fields)
    manual_keys = {item.key for item in definitions if item.measured_key is None}
    selected = [key for key, value in request.fields.items() if value.include]
    manual_values = {
        key: value.value
        for key, value in request.fields.items()
        if key in manual_keys and value.value
    }
    prefix = f"shipping_report/{request.report_type.value}/"
    settings.setValue(prefix + "selected_fields", json.dumps(selected, ensure_ascii=False))
    settings.setValue(prefix + "manual_values", json.dumps(manual_values, ensure_ascii=False))
    settings.setValue(
        prefix + "custom_fields",
        json.dumps([_field_payload(item) for item in request.custom_fields], ensure_ascii=False),
    )
    if request.spectrum_axis is not None:
        settings.setValue(prefix + "axis_mode", request.spectrum_axis.mode.value)
        settings.setValue(prefix + "axis_minimum", request.spectrum_axis.minimum)
        settings.setValue(prefix + "axis_maximum", request.spectrum_axis.maximum)
    settings.sync()


def generate_shipping_report(
    source_path: Path | str,
    output_path: Path | str,
    request: ShippingReportRequest,
) -> Path:
    inspection = inspect_shipping_workbook(source_path)
    validate_shipping_report_request(inspection, request)
    target = Path(output_path).expanduser()
    if target.suffix.lower() != ".pdf":
        target = target.with_suffix(".pdf")
    target = target.resolve()
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f".{target.stem}.tmp.pdf")
    if temporary.exists():
        temporary.unlink()
    try:
        renderer = _ShippingReportRenderer(temporary, inspection, request)
        renderer.render()
        if not temporary.is_file() or temporary.stat().st_size == 0:
            raise RuntimeError("PDF 生成结果为空")
        os.replace(temporary, target)
    finally:
        if temporary.exists():
            temporary.unlink()
    return target


def render_shipping_report_preview(
    source_path: Path | str,
    request: ShippingReportRequest,
    *,
    page_width_px: int = 720,
) -> list[str]:
    """Render the final report layout into compact PNG pages for live preview.

    The preview first uses the same PDF renderer as the saved report, rather
    than maintaining a second HTML approximation.  This keeps page geometry,
    charts, fonts, and the approved header/footer artwork in sync with export.
    The temporary PDF and rendered images stay in a private temporary directory.
    """

    if page_width_px < 240 or page_width_px > 1600:
        raise ValueError("预览宽度必须在 240 到 1600 像素之间")
    inspection = inspect_shipping_workbook(source_path)
    validate_shipping_report_request(inspection, request)
    with tempfile.TemporaryDirectory(prefix="shipping-report-preview-") as directory:
        temporary = Path(directory) / "preview.pdf"
        renderer = _ShippingReportRenderer(temporary, inspection, request)
        renderer.render()
        if not temporary.is_file() or temporary.stat().st_size == 0:
            raise RuntimeError("出货报告预览生成失败")

        document = QPdfDocument()
        pages: list[str] = []
        try:
            if document.load(str(temporary)) != QPdfDocument.Error.None_:
                raise RuntimeError("无法读取出货报告预览")
            if document.pageCount() < 1:
                raise RuntimeError("出货报告预览没有页面")

            for page_number in range(document.pageCount()):
                page_size = document.pagePointSize(page_number)
                if page_size.width() <= 0 or page_size.height() <= 0:
                    raise RuntimeError("出货报告预览页面尺寸无效")
                page_height_px = max(1, round(page_width_px * page_size.height() / page_size.width()))
                image = document.render(page_number, QSize(page_width_px, page_height_px))
                if image.isNull():
                    raise RuntimeError("无法绘制出货报告预览页面")
                data = QByteArray()
                buffer = QBuffer(data)
                try:
                    if not buffer.open(QIODevice.OpenModeFlag.WriteOnly):
                        raise RuntimeError("无法写入出货报告预览")
                    if not image.save(buffer, "PNG"):
                        raise RuntimeError("无法编码出货报告预览")
                finally:
                    buffer.close()
                pages.append(base64.b64encode(bytes(data)).decode("ascii"))
        finally:
            document.close()
            del document
            gc.collect()
        return pages


def _resource_path(relative: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parents[1]))
    return base / relative


def _load_report_font_family(
    paths: Iterable[Path],
    preferred_families: Iterable[str],
) -> str | None:
    """Register a report font and return its preferred Qt family name."""

    preferred = tuple(preferred_families)
    loaded_families: list[str] = []
    for path in paths:
        if not path.is_file():
            continue
        font_id = QFontDatabase.addApplicationFont(str(path))
        if font_id >= 0:
            loaded_families.extend(QFontDatabase.applicationFontFamilies(font_id))

    available = set(QFontDatabase.families())
    for candidate in preferred:
        if candidate in loaded_families or candidate in available:
            return candidate
    return loaded_families[0] if loaded_families else None


def _available_font_families() -> tuple[str, str]:
    """Load the Latin and Chinese fonts used by dynamically drawn report text.

    Some headless Qt Windows platforms expose an empty system font database.  In
    that case QPainter silently falls back to a non-embeddable placeholder font,
    which Poppler renders as square glyphs. Registering concrete font files before
    QPdfWriter starts drawing makes both GUI and headless output deterministic.

    Times New Roman is the primary family, so Latin letters, numbers and symbols
    use it. Microsoft YaHei is the fallback family and supplies all Chinese
    glyphs in mixed-language strings.
    """

    windows_fonts = Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts"
    latin_family = _load_report_font_family(
        (
            windows_fonts / "times.ttf",
            windows_fonts / "timesbd.ttf",
        ),
        ("Times New Roman",),
    )
    chinese_family = _load_report_font_family(
        (
            windows_fonts / "msyh.ttc",
            windows_fonts / "msyhbd.ttc",
            _resource_path("assets/report/report_font.ttf"),
        ),
        ("Microsoft YaHei", "微软雅黑"),
    )
    default_family = QFont().defaultFamily()
    return latin_family or default_family, chinese_family or default_family


def _nice_ceiling(value: float) -> float:
    if value <= 0 or not math.isfinite(value):
        return 1.0
    exponent = math.floor(math.log10(value))
    fraction = value / (10**exponent)
    nice = 1.0 if fraction <= 1 else 2.0 if fraction <= 2 else 5.0 if fraction <= 5 else 10.0
    return nice * (10**exponent)


def _compact_axis_maximum(value: float, *, target_intervals: int = 7) -> tuple[float, float]:
    """Return a readable axis maximum without leaving a large empty tail."""
    if value <= 0 or not math.isfinite(value):
        return 1.0, 0.2
    raw_step = value / max(1, target_intervals)
    exponent = math.floor(math.log10(raw_step))
    scale = 10**exponent
    fraction = raw_step / scale
    step_fraction = min((1.0, 2.0, 2.5, 5.0, 10.0), key=lambda item: abs(item - fraction))
    step = step_fraction * scale
    maximum = math.ceil((value * 1.03) / step) * step
    if maximum <= value:
        maximum += step
    return maximum, step


def _reference_axis_step(span: float, *, target_intervals: int = 5) -> float:
    """Choose a 1/2/2.5/5 scale interval like the approved chart sample."""
    if span <= 0 or not math.isfinite(span):
        return 1.0
    raw_step = span / max(1, target_intervals)
    exponent = math.floor(math.log10(raw_step))
    scale = 10**exponent
    candidates = tuple(multiplier * scale for multiplier in (1.0, 2.0, 2.5, 5.0, 10.0))
    return min(candidates, key=lambda candidate: abs(candidate - raw_step))


def _reference_axis_bounds(
    values: Iterable[float],
    *,
    singleton_span_fraction: float,
    padding_fraction: float = 0.15,
    minimum: float | None = None,
    snap_to_major_ticks: bool = False,
) -> tuple[float, float, float]:
    """Return compact non-zero bounds with five journal-style major intervals.

    The reference supplied for the report intentionally focuses the vertical
    axes around the measured data rather than always forcing a zero baseline.
    A zero lower limit is still retained when it is needed to include a
    non-positive measurement.
    """
    finite_values = [float(value) for value in values if math.isfinite(value)]
    if not finite_values:
        return 0.0, 1.0, 0.2
    lower = min(finite_values)
    upper = max(finite_values)
    if math.isclose(lower, upper, rel_tol=1e-12, abs_tol=1e-12):
        span = max(abs(lower) * singleton_span_fraction, 1.0)
        lower -= span / 2.0
        upper += span / 2.0
    else:
        padding = (upper - lower) * padding_fraction
        lower -= padding
        upper += padding
    if minimum is not None:
        lower = max(minimum, lower)
    if upper <= lower:
        upper = lower + max(abs(lower) * singleton_span_fraction, 1.0)
    major_step = _reference_axis_step(upper - lower)
    if snap_to_major_ticks:
        lower = math.floor(lower / major_step) * major_step
        upper = math.ceil(upper / major_step) * major_step
        if upper <= lower:
            upper = lower + major_step
    return lower, upper, major_step


def _axis_tick_values(minimum: float, maximum: float, step: float) -> list[float]:
    """Return stable tick values inside a floating-point axis interval."""
    if step <= 0 or not math.isfinite(step):
        return []
    first = math.ceil(minimum / step - 1e-9)
    last = math.floor(maximum / step + 1e-9)
    return [index * step for index in range(first, last + 1)]


def _format_axis_tick(value: float, step: float, *, force_decimal: bool = False) -> str:
    """Keep tick labels compact while preserving the sample's decimal style."""
    decimals = max(0, int(math.ceil(-math.log10(abs(step)) - 1e-9))) if step else 0
    if force_decimal:
        decimals = max(1, decimals)
    return f"{value:.{decimals}f}"


def _format_pole_summary_value(definition: ReportFieldDefinition, value: str) -> str:
    """Render Pole summary values at a compact, report-appropriate precision."""

    precision = {
        "power": 1,
        "operatingCurrent": 2,
        "operatingVoltage": 2,
        "slopeEfficiency": 2,
        "electroOpticalEfficiency": 2,
        "coolingTemperature": 1,
    }.get(definition.key)
    if precision is None:
        return value

    numeric_value = _finite_or_none(value)
    if numeric_value is None:
        return value
    if math.isclose(numeric_value, 0.0, abs_tol=10 ** -(precision + 1)):
        numeric_value = 0.0
    return f"{numeric_value:.{precision}f}".rstrip("0").rstrip(".") or "0"


def _format_performance_parameters_title(operating_current_a: float) -> str:
    """Return the English performance-summary heading for the selected current."""

    return f"Performance Parameters @{operating_current_a:g}A"


class _ShippingReportRenderer:
    A4_WIDTH_MM = 210.0
    A4_HEIGHT_MM = 297.0
    FOOTER_TOP_MM = 258.0
    HEADER_ARTWORK_TOP_MM = 5.0
    TITLE_TOP_MM = 32.5
    SERIAL_NUMBER_TOP_MM = 44.5
    TITLE_DIVIDER_Y_MM = 57.0
    POLE_SUMMARY_TOP_MM = 62.0
    POLE_TABLE_VALUE_FONT_SIZE = 10.7
    DETAIL_TABLE_VALUE_FONT_SIZE = POLE_TABLE_VALUE_FONT_SIZE

    def __init__(self, path: Path, inspection: WorkbookInspection, request: ShippingReportRequest) -> None:
        global _REPORT_QT_APPLICATION
        application = QApplication.instance()
        if application is None:
            _REPORT_QT_APPLICATION = QApplication([])
            application = _REPORT_QT_APPLICATION
        self._application = application
        self.path = path
        self.inspection = inspection
        self.request = request
        self.latin_font_family, self.chinese_font_family = _available_font_families()
        self.writer = QPdfWriter(str(path))
        self.writer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
        self.writer.setPageMargins(QMarginsF(0, 0, 0, 0), QPageLayout.Unit.Millimeter)
        self.writer.setResolution(144)
        self.writer.setTitle(f"{request.product_name} - {request.sn} Shipping Report")
        self.writer.setCreator("Pump Driver Integrated Test")
        self.painter = QPainter()
        self.page_number = 0
        self.scale = self.writer.resolution() / 25.4

    def mm(self, value: float) -> float:
        return value * self.scale

    def rect(self, x: float, y: float, width: float, height: float) -> QRectF:
        return QRectF(self.mm(x), self.mm(y), self.mm(width), self.mm(height))

    def font(self, size: float, *, bold: bool = False) -> QFont:
        result = QFont(self.latin_font_family)
        result.setFamilies((self.latin_font_family, self.chinese_font_family))
        result.setPointSizeF(size)
        result.setBold(bold)
        return result

    def _reference_image(self, relative_path: str) -> QImage:
        """Load approved report artwork without recreating any of its content."""

        path = _resource_path(relative_path)
        image = QImage(str(path))
        if image.isNull():
            raise RuntimeError(f"无法读取出货报告原始图片：{path}")
        return image

    def render(self) -> None:
        if not self.painter.begin(self.writer):
            raise RuntimeError("无法启动 PDF 绘制")
        try:
            self.painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
            self.painter.setRenderHint(QPainter.RenderHint.TextAntialiasing, True)
            if self.request.report_type is ShippingReportType.SPECTRUM:
                self._render_spectrum_report()
            else:
                self._render_pole_report()
        finally:
            self.painter.end()

    def _new_page(self, continuation_title: str = "") -> None:
        if self.page_number:
            self.writer.newPage()
        self.page_number += 1
        self._draw_header()
        self._draw_title(continuation_title)
        self._draw_footer()

    def _draw_header(self) -> None:
        # This is the user-approved complete header.  Keep it as one raster so
        # the logo, wording, font weights and spacing cannot drift independently.
        header = self._reference_image("assets/report/everbright_header_reference.png")
        header_height_mm = self.A4_WIDTH_MM * header.height() / header.width()
        self.painter.drawImage(
            self.rect(0, self.HEADER_ARTWORK_TOP_MM, self.A4_WIDTH_MM, header_height_mm),
            header,
        )

    def _draw_title(self, continuation_title: str = "") -> None:
        self.painter.setPen(QColor("#090909"))
        self.painter.setFont(self.font(16, bold=True))
        title = self.request.product_name if not continuation_title else f"{self.request.product_name} - {continuation_title}"
        self.painter.drawText(
            self.rect(20, self.TITLE_TOP_MM, 170, 10),
            int(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter),
            title,
        )
        self.painter.setFont(self.font(16, bold=True))
        self.painter.drawText(
            self.rect(20, self.SERIAL_NUMBER_TOP_MM, 170, 8),
            int(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter),
            f"TEST SN: {self.request.sn}",
        )
        self.painter.setPen(QPen(QColor("#111111"), self.mm(0.35)))
        self.painter.drawLine(
            QPointF(self.mm(19), self.mm(self.TITLE_DIVIDER_Y_MM)),
            QPointF(self.mm(191), self.mm(self.TITLE_DIVIDER_Y_MM)),
        )

    def _draw_footer(self) -> None:
        top = self.FOOTER_TOP_MM
        self.painter.setPen(QPen(QColor("#d1d1d1"), self.mm(0.25)))
        self.painter.drawLine(QPointF(self.mm(18), self.mm(top)), QPointF(self.mm(192), self.mm(top)))
        self.painter.setPen(QColor("#1a1a1a"))
        self.painter.setFont(self.font(6.7))
        notice = (
            "通告：长光华芯光电不断提高产品，为我们的客户提供卓越的品质和可靠性。"
            "在任何时候我们可能会更改规格和产品说明，不另行通知。此外，我们提供了保质期以确保客户的满意度。"
            "如需完整的详细资讯，请联系我们的销售代表。"
        )
        self.painter.drawText(
            self.rect(12, top + 5, 75, 28),
            int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop | Qt.TextFlag.TextWordWrap),
            notice,
        )
        # The warning artwork includes all of its original text.  Do not overlay
        # translated/retyped content or stretch it away from its source ratio.
        warning = self._reference_image("assets/report/laser_warning_reference.png")
        warning_height_mm = 27.0
        warning_width_mm = warning_height_mm * warning.width() / warning.height()
        warning_x_mm = 104.5 - warning_width_mm / 2
        self.painter.drawImage(
            self.rect(warning_x_mm, top + 4, warning_width_mm, warning_height_mm),
            warning,
        )
        self.painter.setFont(self.font(7.1))
        company = (
            "苏州长光华芯光电技术股份有限公司\n"
            "苏州高新区漓江路 56 号\n"
            "电话：0512-69372570\n传真：0512-69372559\n"
            "版权所有 长光华芯光电技术股份有限公司"
        )
        self.painter.drawText(
            self.rect(122, top + 5, 76, 28),
            int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop),
            company,
        )
        self.painter.setFont(self.font(6.3))
        self.painter.setPen(QColor("#666666"))
        self.painter.drawText(
            self.rect(90, 291, 30, 4),
            int(Qt.AlignmentFlag.AlignCenter),
            f"第 {self.page_number} 页",
        )

    def _selected_fields(self, definitions: Iterable[ReportFieldDefinition]) -> list[tuple[ReportFieldDefinition, str]]:
        output = []
        for definition in definitions:
            selected = self.request.fields.get(definition.key, SelectedReportField(False, ""))
            if selected.include:
                output.append((definition, selected.value))
        return output

    def _definitions(self, report_type: ShippingReportType) -> tuple[ReportFieldDefinition, ...]:
        return _report_definitions(report_type, self.request.custom_fields)

    def _render_spectrum_report(self) -> None:
        self._new_page()
        fields = self._selected_fields(self._definitions(ShippingReportType.SPECTRUM))
        left = [item for item in fields if item[0].side == "left"]
        right = [item for item in fields if item[0].side == "right"]
        summary_height = self._draw_spectrum_summary(72, left, right)
        chart_top = min(116.0, 75.0 + summary_height + 4.0)
        available = self.FOOTER_TOP_MM - chart_top - 7.0
        chart_height = max(50.0, available / 2.0 - 2.0)
        self._draw_power_efficiency_chart(self.rect(23, chart_top, 164, chart_height), title="")
        self._draw_spectrum_chart(self.rect(23, chart_top + chart_height + 4, 164, chart_height))

    def _draw_spectrum_summary(
        self,
        top: float,
        left: list[tuple[ReportFieldDefinition, str]],
        right: list[tuple[ReportFieldDefinition, str]],
    ) -> float:
        row_count = max(len(left), len(right), 1)
        row_height = min(6.1, 36.0 / row_count)
        height = row_count * row_height
        x, width = 21.0, 168.0
        self.painter.setPen(QPen(QColor("#333333"), self.mm(0.2)))
        self.painter.drawLine(QPointF(self.mm(x), self.mm(top)), QPointF(self.mm(x + width), self.mm(top)))
        self.painter.drawLine(QPointF(self.mm(x), self.mm(top + height)), QPointF(self.mm(x + width), self.mm(top + height)))
        self.painter.drawLine(QPointF(self.mm(105), self.mm(top)), QPointF(self.mm(105), self.mm(top + height)))
        for side_x, values in ((23.0, left), (107.0, right)):
            for index, (definition, value) in enumerate(values):
                y = top + index * row_height
                self.painter.setPen(QColor("#161616"))
                self.painter.setFont(self.font(8.5, bold=True))
                self.painter.drawText(self.rect(side_x, y, 43, row_height), int(Qt.AlignmentFlag.AlignVCenter), definition.label)
                self.painter.setFont(self.font(8.6, bold=True))
                self.painter.drawText(self.rect(side_x + 47, y, 20, row_height), int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter), value)
                self.painter.setFont(self.font(8.2))
                self.painter.drawText(self.rect(side_x + 69, y, 13, row_height), int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter), definition.unit)
        return height

    def _render_pole_report(self) -> None:
        fields = self._selected_fields(self._definitions(ShippingReportType.POLE))
        points = list(self.inspection.points)
        first_capacity = 12
        remaining = points
        self._new_page()
        summary_height = self._draw_pole_summary(self.POLE_SUMMARY_TOP_MM, fields)
        chart_top = self.POLE_SUMMARY_TOP_MM + summary_height + 6.0
        chart_height = 62.0
        self._draw_power_efficiency_chart(self.rect(31, chart_top, 148, chart_height), title="")
        table_top = chart_top + chart_height
        room_rows = max(1, min(first_capacity, int((self.FOOTER_TOP_MM - table_top - 5.0) / 6.0) - 1))
        current_page = remaining[:room_rows]
        remaining = remaining[room_rows:]
        self._draw_detail_table(table_top, current_page)
        while remaining:
            self._new_page("Detailed Measurement Data (Continued)")
            capacity = max(1, int((self.FOOTER_TOP_MM - 76.0) / 6.0) - 1)
            current_page = remaining[:capacity]
            remaining = remaining[capacity:]
            self._draw_detail_table(74.0, current_page)

    def _draw_pole_summary(self, top: float, fields: list[tuple[ReportFieldDefinition, str]]) -> float:
        title_height = 8.0
        title_gap = 3.0
        header_height = 7.2
        row_height = 7.4
        table_x, table_width = 18.0, 174.0
        parameter_width = table_width * 0.45
        value_width = table_width * 0.30
        unit_width = table_width - parameter_width - value_width
        cell_padding = 3.5
        divider_pen = QPen(QColor("#dfe4e8"), self.mm(0.18))

        self._draw_section_title(
            top,
            _format_performance_parameters_title(self.request.operating_current_a),
            height=title_height,
            font_size=14.0,
            accent_width=0.75,
            accent_height=5.0,
        )
        table_top = top + title_height + title_gap
        table_height = header_height + len(fields) * row_height

        # A lightly tinted header and horizontal dividers establish the columns
        # without the heavy grid or alternating body fills of the old layout.
        self.painter.fillRect(self.rect(table_x, table_top, table_width, header_height), QColor("#f6f8fa"))
        self.painter.setPen(divider_pen)
        self.painter.drawLine(
            QPointF(self.mm(table_x), self.mm(table_top)),
            QPointF(self.mm(table_x + table_width), self.mm(table_top)),
        )
        self.painter.drawLine(
            QPointF(self.mm(table_x), self.mm(table_top + header_height)),
            QPointF(self.mm(table_x + table_width), self.mm(table_top + header_height)),
        )

        self.painter.setPen(QColor("#0b4b82"))
        self.painter.setFont(self.font(10.3, bold=True))
        self.painter.drawText(
            self.rect(table_x + cell_padding, table_top, parameter_width - 2 * cell_padding, header_height),
            int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
            "Parameter",
        )
        self.painter.drawText(
            self.rect(
                table_x + parameter_width + cell_padding,
                table_top,
                value_width - 2 * cell_padding,
                header_height,
            ),
            int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter),
            "Value",
        )
        self.painter.drawText(
            self.rect(
                table_x + parameter_width + value_width + cell_padding,
                table_top,
                unit_width - 2 * cell_padding,
                header_height,
            ),
            int(Qt.AlignmentFlag.AlignCenter),
            "Unit",
        )

        for index, (definition, value) in enumerate(fields):
            y = table_top + header_height + index * row_height
            self.painter.setPen(QColor("#252a2e"))
            self.painter.setFont(self.font(self.POLE_TABLE_VALUE_FONT_SIZE))
            self.painter.drawText(
                self.rect(table_x + cell_padding, y, parameter_width - 2 * cell_padding, row_height),
                int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
                definition.label,
            )
            self.painter.drawText(
                self.rect(
                    table_x + parameter_width + cell_padding,
                    y,
                    value_width - 2 * cell_padding,
                    row_height,
                ),
                int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter),
                _format_pole_summary_value(definition, value),
            )
            self.painter.setFont(self.font(10.5))
            self.painter.drawText(
                self.rect(
                    table_x + parameter_width + value_width + cell_padding,
                    y,
                    unit_width - 2 * cell_padding,
                    row_height,
                ),
                int(Qt.AlignmentFlag.AlignCenter),
                definition.unit,
            )
            self.painter.setPen(divider_pen)
            self.painter.drawLine(
                QPointF(self.mm(table_x), self.mm(y + row_height)),
                QPointF(self.mm(table_x + table_width), self.mm(y + row_height)),
            )
        return title_height + title_gap + table_height

    def _draw_power_efficiency_chart(self, outer: QRectF, *, title: str) -> None:
        points = self.inspection.points
        currents = [point.current_a for point in points]
        powers = [point.power_w for point in points]
        efficiencies = [point.efficiency * 100.0 for point in points]

        def chart_font(size: float, *, bold: bool = False) -> QFont:
            """Use the clean sans-serif typography of the report chart standard."""
            result = QFont("Arial")
            result.setFamilies(("Arial", "Helvetica", self.chinese_font_family))
            result.setPointSizeF(size)
            result.setBold(bold)
            return result

        left_margin = self.mm(17)
        right_margin = self.mm(18)
        top_margin = self.mm(10 if title else 3)
        bottom_margin = self.mm(12)
        plot = QRectF(
            outer.left() + left_margin,
            outer.top() + top_margin,
            outer.width() - left_margin - right_margin,
            outer.height() - top_margin - bottom_margin,
        )
        if title:
            self.painter.setPen(QColor("#343434"))
            self.painter.setFont(chart_font(10.2, bold=True))
            self.painter.drawText(
                QRectF(outer.left(), outer.top(), outer.width(), self.mm(8)),
                int(Qt.AlignmentFlag.AlignCenter),
                title,
            )
        # Default shipment-report style: retain a compact dual-Y layout and
        # label the measured current points directly instead of adding a
        # synthetic zero-current tick when it was not measured.
        x_data_min = min(currents)
        x_data_max = max(currents)
        if math.isclose(x_data_min, x_data_max, rel_tol=1e-12, abs_tol=1e-12):
            x_step = _reference_axis_step(max(abs(x_data_min), 1.0), target_intervals=4)
            x_min = x_data_min - 0.75 * x_step
            x_max = x_data_max + 0.75 * x_step
        else:
            unique_currents = sorted(set(currents))
            measured_steps = [
                current - previous
                for previous, current in zip(unique_currents, unique_currents[1:])
                if current > previous
            ]
            x_step = _reference_axis_step(min(measured_steps), target_intervals=1)
            x_padding = max(0.25 * x_step, 0.04 * (x_data_max - x_data_min))
            x_min = x_data_min - x_padding
            x_max = x_data_max + x_padding
        power_min, power_max, power_step = _reference_axis_bounds(
            powers,
            singleton_span_fraction=1.2,
            minimum=0.0,
        )
        efficiency_min, efficiency_max, efficiency_step = _reference_axis_bounds(
            efficiencies,
            singleton_span_fraction=0.5,
            minimum=0.0,
            snap_to_major_ticks=True,
        )

        def map_x(value: float) -> float:
            return plot.left() + (value - x_min) / (x_max - x_min) * plot.width()

        def map_power(value: float) -> float:
            return plot.bottom() - (value - power_min) / (power_max - power_min) * plot.height()

        def map_efficiency(value: float) -> float:
            return plot.bottom() - (value - efficiency_min) / (efficiency_max - efficiency_min) * plot.height()

        axis_color = QColor("#202020")
        power_color = QColor("#0072B2")
        efficiency_label_color = QColor("#D55E00")
        major_tick_pen = QPen(power_color, self.mm(0.32))
        minor_tick_pen = QPen(power_color, self.mm(0.22))
        right_major_tick_pen = QPen(efficiency_label_color, self.mm(0.32))
        right_minor_tick_pen = QPen(efficiency_label_color, self.mm(0.22))
        x_major_tick_pen = QPen(axis_color, self.mm(0.32))
        x_minor_tick_pen = QPen(axis_color, self.mm(0.22))
        major_tick_length = self.mm(1.55)
        minor_tick_length = self.mm(0.85)

        def is_major_tick(value: float, major_values: list[float]) -> bool:
            return any(math.isclose(value, major, rel_tol=0.0, abs_tol=1e-8) for major in major_values)

        power_ticks = _axis_tick_values(power_min, power_max, power_step)
        power_minor_ticks = _axis_tick_values(power_min, power_max, power_step / 2.0)
        self.painter.setFont(chart_font(8.0))
        for value in power_minor_ticks:
            if is_major_tick(value, power_ticks):
                continue
            y = map_power(value)
            self.painter.setPen(minor_tick_pen)
            self.painter.drawLine(QPointF(plot.left(), y), QPointF(plot.left() + minor_tick_length, y))
        for value in power_ticks:
            y = map_power(value)
            self.painter.setPen(major_tick_pen)
            self.painter.drawLine(QPointF(plot.left(), y), QPointF(plot.left() + major_tick_length, y))
            self.painter.setPen(power_color)
            self.painter.drawText(
                QRectF(plot.left() - self.mm(14), y - self.mm(2.2), self.mm(12), self.mm(4.4)),
                int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter),
                _format_axis_tick(value, power_step),
            )

        efficiency_ticks = _axis_tick_values(efficiency_min, efficiency_max, efficiency_step)
        efficiency_minor_ticks = _axis_tick_values(efficiency_min, efficiency_max, efficiency_step / 2.0)
        for value in efficiency_minor_ticks:
            if is_major_tick(value, efficiency_ticks):
                continue
            y = map_efficiency(value)
            self.painter.setPen(right_minor_tick_pen)
            self.painter.drawLine(QPointF(plot.right(), y), QPointF(plot.right() - minor_tick_length, y))
        for value in efficiency_ticks:
            y = map_efficiency(value)
            self.painter.setPen(right_major_tick_pen)
            self.painter.drawLine(QPointF(plot.right(), y), QPointF(plot.right() - major_tick_length, y))
            self.painter.setPen(efficiency_label_color)
            self.painter.drawText(
                QRectF(plot.right() + self.mm(2), y - self.mm(2.2), self.mm(13), self.mm(4.4)),
                int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
                _format_axis_tick(value, efficiency_step),
            )

        x_ticks = _axis_tick_values(x_min, x_max, x_step)
        x_minor_ticks = _axis_tick_values(x_min, x_max, x_step / 2.0)
        for value in x_minor_ticks:
            if is_major_tick(value, x_ticks):
                continue
            x = map_x(value)
            self.painter.setPen(x_minor_tick_pen)
            self.painter.drawLine(QPointF(x, plot.bottom()), QPointF(x, plot.bottom() - minor_tick_length))
            self.painter.drawLine(QPointF(x, plot.top()), QPointF(x, plot.top() + minor_tick_length))
        for value in x_ticks:
            x = map_x(value)
            self.painter.setPen(x_major_tick_pen)
            self.painter.drawLine(QPointF(x, plot.bottom()), QPointF(x, plot.bottom() - major_tick_length))
            self.painter.drawLine(QPointF(x, plot.top()), QPointF(x, plot.top() + major_tick_length))
            self.painter.setPen(axis_color)
            self.painter.drawText(
                QRectF(x - self.mm(7), plot.bottom() + self.mm(1.5), self.mm(14), self.mm(4.5)),
                int(Qt.AlignmentFlag.AlignCenter),
                _format_axis_tick(value, x_step),
            )

        # Frame and axis labels mirror the standard shipment-report reference:
        # neutral horizontal frame, blue left power axis, and orange right axis.
        frame_pen = QPen(axis_color, self.mm(0.42))
        self.painter.setPen(frame_pen)
        self.painter.drawLine(QPointF(plot.left(), plot.top()), QPointF(plot.right(), plot.top()))
        self.painter.drawLine(QPointF(plot.left(), plot.bottom()), QPointF(plot.right(), plot.bottom()))
        self.painter.setPen(QPen(power_color, self.mm(0.42)))
        self.painter.drawLine(QPointF(plot.left(), plot.top()), QPointF(plot.left(), plot.bottom()))
        self.painter.setPen(QPen(efficiency_label_color, self.mm(0.42)))
        self.painter.drawLine(QPointF(plot.right(), plot.top()), QPointF(plot.right(), plot.bottom()))

        self.painter.setFont(chart_font(8.8))
        self.painter.setPen(axis_color)
        self.painter.drawText(
            QRectF(plot.left(), plot.bottom() + self.mm(6.5), plot.width(), self.mm(5.5)),
            int(Qt.AlignmentFlag.AlignCenter),
            "Drive current (A)",
        )
        vertical_label_span = outer.height() - self.mm(2.0)
        vertical_label_height = self.mm(6.5)
        self.painter.save()
        self.painter.translate(outer.left() + self.mm(3.2), plot.center().y())
        self.painter.rotate(-90)
        self.painter.setPen(power_color)
        self.painter.drawText(
            QRectF(-vertical_label_span / 2, -vertical_label_height / 2, vertical_label_span, vertical_label_height),
            int(Qt.AlignmentFlag.AlignCenter),
            "Output power (W)",
        )
        self.painter.restore()
        self.painter.save()
        self.painter.translate(outer.right() - self.mm(3.2), plot.center().y())
        self.painter.rotate(90)
        self.painter.setPen(efficiency_label_color)
        self.painter.drawText(
            QRectF(-vertical_label_span / 2, -vertical_label_height / 2, vertical_label_span, vertical_label_height),
            int(Qt.AlignmentFlag.AlignCenter),
            "E-O efficiency (%)",
        )
        self.painter.restore()

        power_points = QPolygonF([QPointF(map_x(x), map_power(y)) for x, y in zip(currents, powers)])
        efficiency_points = QPolygonF([QPointF(map_x(x), map_efficiency(y)) for x, y in zip(currents, efficiencies)])
        marker_outline = QPen(QColor("#ffffff"), self.mm(0.24))
        self.painter.setPen(QPen(power_color, self.mm(0.60)))
        self.painter.drawPolyline(power_points)
        self.painter.setBrush(power_color)
        self.painter.setPen(marker_outline)
        for point in power_points:
            self.painter.drawEllipse(point, self.mm(1.15), self.mm(1.15))
        self.painter.setPen(QPen(efficiency_label_color, self.mm(0.60)))
        self.painter.drawPolyline(efficiency_points)
        self.painter.setBrush(efficiency_label_color)
        self.painter.setPen(marker_outline)
        for point in efficiency_points:
            self.painter.drawRect(QRectF(point.x() - self.mm(1.05), point.y() - self.mm(1.05), self.mm(2.1), self.mm(2.1)))

        # Keep the legend away from the fast-rising efficiency curve.  The
        # lower-right corner is normally clear for LIV data and matches the
        # report's approved chart layout.
        legend_x = plot.right() - self.mm(55.0)
        legend_y = plot.bottom() - self.mm(10.5)
        legend_line_width = self.mm(11.0)
        legend_label_x = legend_x + self.mm(13.0)
        self.painter.setFont(chart_font(8.0))

        def draw_legend_entry(y: float, color: QColor, label: str, marker: str) -> None:
            line_end = legend_x + legend_line_width
            marker_center = (legend_x + line_end) / 2
            self.painter.setPen(QPen(color, self.mm(0.56)))
            self.painter.drawLine(QPointF(legend_x, y), QPointF(line_end, y))
            self.painter.setBrush(color)
            self.painter.setPen(marker_outline)
            if marker == "circle":
                self.painter.drawEllipse(QPointF(marker_center, y), self.mm(1.05), self.mm(1.05))
            else:
                self.painter.drawRect(
                    QRectF(marker_center - self.mm(1.0), y - self.mm(1.0), self.mm(2.0), self.mm(2.0))
                )
            self.painter.setPen(axis_color)
            self.painter.drawText(
                QRectF(legend_label_x, y - self.mm(2.5), plot.width() - (legend_label_x - plot.left()) - self.mm(2), self.mm(5.0)),
                int(Qt.AlignmentFlag.AlignVCenter),
                label,
            )

        draw_legend_entry(legend_y, power_color, "Output power", "circle")
        draw_legend_entry(legend_y + self.mm(6.3), efficiency_label_color, "E-O efficiency", "square")
        self.painter.setBrush(Qt.BrushStyle.NoBrush)

    def _draw_spectrum_chart(self, outer: QRectF) -> None:
        axis = self.request.spectrum_axis
        assert axis is not None
        legend_width = self.mm(22)
        plot = QRectF(
            outer.left() + self.mm(16),
            outer.top() + self.mm(10),
            outer.width() - self.mm(16) - self.mm(10) - legend_width,
            outer.height() - self.mm(10) - self.mm(12),
        )
        self.painter.setPen(QColor("#343434"))
        self.painter.setFont(self.font(10.2, bold=True))
        self.painter.drawText(
            QRectF(outer.left(), outer.top(), outer.width(), self.mm(8)),
            int(Qt.AlignmentFlag.AlignCenter),
            "Spectrum at Different Currents",
        )
        x_values = [value for point in self.inspection.points for value in point.wavelength]
        x_min, x_max = min(x_values), max(x_values)
        if math.isclose(x_min, x_max):
            x_max = x_min + 1.0

        def map_x(value: float) -> float:
            return plot.left() + (value - x_min) / (x_max - x_min) * plot.width()

        def map_y(value: float) -> float:
            return plot.bottom() - (value - axis.minimum) / (axis.maximum - axis.minimum) * plot.height()

        self.painter.setFont(self.font(6.7))
        for index in range(5):
            ratio = index / 4
            y = plot.bottom() - ratio * plot.height()
            value = axis.minimum + ratio * (axis.maximum - axis.minimum)
            self.painter.setPen(QPen(QColor("#d7dade"), self.mm(0.13)))
            self.painter.drawLine(QPointF(plot.left(), y), QPointF(plot.right(), y))
            self.painter.setPen(QColor("#333333"))
            self.painter.drawText(QRectF(plot.left() - self.mm(13), y - self.mm(2), self.mm(11), self.mm(4)), int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter), f"{value:g}")
        for index in range(5):
            ratio = index / 4
            x = plot.left() + ratio * plot.width()
            self.painter.setPen(QPen(QColor("#e0e3e6"), self.mm(0.12)))
            self.painter.drawLine(QPointF(x, plot.top()), QPointF(x, plot.bottom()))
            self.painter.setPen(QColor("#333333"))
            self.painter.drawText(QRectF(x - self.mm(8), plot.bottom() + self.mm(1.5), self.mm(16), self.mm(4)), int(Qt.AlignmentFlag.AlignCenter), f"{x_min + ratio * (x_max - x_min):.1f}")
        self.painter.setPen(QPen(QColor("#343434"), self.mm(0.25)))
        self.painter.drawRect(plot)
        self.painter.setFont(self.font(7.4, bold=True))
        self.painter.drawText(
            QRectF(plot.left(), plot.bottom() + self.mm(6.5), plot.width(), self.mm(5)),
            int(Qt.AlignmentFlag.AlignCenter),
            "Wavelength (nm)",
        )
        self.painter.save()
        self.painter.translate(outer.left() + self.mm(3), plot.center().y())
        self.painter.rotate(-90)
        y_label = "Intensity (counts)" if axis.mode is SpectrumAxisMode.COUNTS else "Relative Intensity (dB)"
        self.painter.drawText(QRectF(-self.mm(30), -self.mm(3), self.mm(60), self.mm(6)), int(Qt.AlignmentFlag.AlignCenter), y_label)
        self.painter.restore()
        colors = (
            "#376f9f", "#c54e4e", "#799442", "#72548c", "#3f9298", "#d27b36",
            "#254f69", "#8b493d", "#536d2d", "#60456f", "#32767a", "#9b5c22",
        )
        self.painter.save()
        self.painter.setClipRect(plot)
        for index, point in enumerate(self.inspection.points):
            values = list(point.intensity)
            if axis.mode is SpectrumAxisMode.RELATIVE_DB:
                peak = max(values)
                if peak <= 0:
                    transformed = [axis.minimum for _value in values]
                else:
                    floor = peak * 1e-12
                    transformed = [10.0 * math.log10(max(value, floor) / peak) for value in values]
            else:
                transformed = values
            stride = max(1, len(values) // 800)
            polygon = QPolygonF()
            for x, y in zip(point.wavelength[::stride], transformed[::stride]):
                if math.isfinite(x) and math.isfinite(y):
                    polygon.append(QPointF(map_x(x), map_y(y)))
            self.painter.setPen(QPen(QColor(colors[index % len(colors)]), self.mm(0.35)))
            self.painter.drawPolyline(polygon)
        self.painter.restore()
        legend_x = plot.right() + self.mm(4)
        legend_y = plot.top()
        legend_row = min(self.mm(4.6), plot.height() / max(1, len(self.inspection.points)))
        self.painter.setFont(self.font(6.3))
        for index, point in enumerate(self.inspection.points):
            y = legend_y + index * legend_row + legend_row / 2
            self.painter.setPen(QPen(QColor(colors[index % len(colors)]), self.mm(0.4)))
            self.painter.drawLine(QPointF(legend_x, y), QPointF(legend_x + self.mm(7), y))
            self.painter.setPen(QColor("#333333"))
            self.painter.drawText(QRectF(legend_x + self.mm(8), y - legend_row / 2, self.mm(13), legend_row), int(Qt.AlignmentFlag.AlignVCenter), f"{point.current_a:g}A")

    def _draw_detail_table(self, top: float, points: list[ReportPoint]) -> None:
        self._draw_section_title(top, "Detailed Measurement Data")
        table_top = top + 9
        row_height = 6.0
        column_widths = (43.5, 43.5, 43.5, 43.5)
        headers = ("Current (A)", "Power (W)", "Voltage (V)", "E-O efficiency (%)")
        x = 18.0
        self.painter.fillRect(self.rect(x, table_top, sum(column_widths), row_height), QColor("#0b4b82"))
        self.painter.setPen(QColor("#ffffff"))
        self.painter.setFont(self.font(7.7, bold=True))
        cursor = x
        for width, header in zip(column_widths, headers):
            self.painter.drawText(self.rect(cursor, table_top, width, row_height), int(Qt.AlignmentFlag.AlignCenter), header)
            cursor += width
        self.painter.setFont(self.font(self.DETAIL_TABLE_VALUE_FONT_SIZE))
        for row_index, point in enumerate(points, start=1):
            y = table_top + row_index * row_height
            self.painter.fillRect(self.rect(x, y, sum(column_widths), row_height), QColor("#f3f4f5" if row_index % 2 else "#ffffff"))
            values = (
                f"{point.current_a:g}",
                f"{point.power_w:.1f}",
                f"{point.voltage_v:.2f}",
                f"{point.efficiency * 100.0:.2f}%",
            )
            cursor = x
            for width, value in zip(column_widths, values):
                self.painter.setPen(QColor("#d0d4d8"))
                self.painter.drawRect(self.rect(cursor, y, width, row_height))
                self.painter.setPen(QColor("#202020"))
                self.painter.drawText(self.rect(cursor, y, width, row_height), int(Qt.AlignmentFlag.AlignCenter), value)
                cursor += width

    def _draw_section_title(
        self,
        top: float,
        title: str,
        *,
        height: float = 7.0,
        font_size: float = 10.5,
        accent_width: float = 1.4,
        accent_height: float | None = None,
    ) -> None:
        accent_height = height if accent_height is None else min(accent_height, height)
        accent_top = top + (height - accent_height) / 2
        self.painter.setPen(QColor("#0b4b82"))
        self.painter.fillRect(self.rect(17, accent_top, accent_width, accent_height), QColor("#0b4b82"))
        self.painter.setFont(self.font(font_size, bold=True))
        self.painter.drawText(self.rect(20, top, 100, height), int(Qt.AlignmentFlag.AlignVCenter), title)
