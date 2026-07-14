"""Excel workbook export for combined test sessions."""

from __future__ import annotations

import math
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet


RESULT_HEADERS = (
    "电流(A)",
    "电压(V)",
    "功率(W)",
    "电光效率",
    "中心波长(nm)",
    "质心波长(nm)",
    "FWHM(nm)",
    "PIB",
    "SMSR(dB)",
)
INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


@dataclass(frozen=True)
class ExcelTestRecord:
    current_a: float
    voltage_v: float
    power_w: float
    efficiency: float
    peak_wavelength_nm: float
    centroid_nm: float
    fwhm_nm: float
    pib: float
    wavelength: Iterable[float]
    intensity: Iterable[float]
    smsr_db: float = math.nan


def sanitize_sn(sn: str) -> str:
    cleaned = INVALID_FILENAME_CHARS.sub("_", sn.strip()).rstrip(". ")
    if not cleaned:
        raise ValueError("SN 不能为空")
    return cleaned


def build_test_workbook_path(output_dir: Path, sn: str, test_time: datetime) -> Path:
    # Distinct sessions must never silently replace each other, even when two
    # tests start during the same second.
    filename = f"{sanitize_sn(sn)}_{test_time.strftime('%Y_%m_%d_%H_%M_%S_%f')}.xlsx"
    return Path(output_dir).expanduser() / filename


def _finite_or_none(value: float) -> float | None:
    number = float(value)
    return number if math.isfinite(number) else None


def _ensure_result_headers(sheet: Worksheet) -> None:
    for column, header in enumerate(RESULT_HEADERS, start=1):
        cell = sheet.cell(row=2, column=column, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.column_dimensions["I"].width = 12


def _create_workbook(path: Path) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试数据"
    sheet["A1"] = "LIV"
    sheet["J1"] = "光谱"
    _ensure_result_headers(sheet)

    sheet.freeze_panes = "A3"
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12
    sheet.column_dimensions["E"].width = 16
    sheet.column_dimensions["F"].width = 16
    sheet.column_dimensions["G"].width = 12
    sheet.column_dimensions["H"].width = 12
    sheet["A1"].font = Font(bold=True)
    sheet["J1"].font = Font(bold=True)
    return workbook


def save_test_records(path: Path, records: Iterable[ExcelTestRecord]) -> None:
    """Write a complete, current-sorted test workbook with a single XLSX save."""
    records_by_current: dict[float, tuple[ExcelTestRecord, list[float], list[float]]] = {}
    for record in records:
        wavelength_values = [float(value) for value in record.wavelength]
        intensity_values = [float(value) for value in record.intensity]
        if len(wavelength_values) != len(intensity_values):
            raise ValueError("波长和强度数据长度必须一致")
        records_by_current[float(record.current_a)] = (record, wavelength_values, intensity_values)
    if not records_by_current:
        raise ValueError("至少需要一个测试记录")

    target = Path(path).expanduser()
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = _create_workbook(target)
    sheet = workbook[workbook.sheetnames[0]]

    spectrum_index = 0
    for index, current in enumerate(sorted(records_by_current)):
        record, wavelength_values, intensity_values = records_by_current[current]
        result_row = 3 + index
        result_values = (
            record.current_a,
            record.voltage_v,
            record.power_w,
            record.efficiency,
            record.peak_wavelength_nm,
            record.centroid_nm,
            record.fwhm_nm,
            record.pib,
            record.smsr_db,
        )
        for column, value in enumerate(result_values, start=1):
            sheet.cell(row=result_row, column=column, value=_finite_or_none(value))
        for column in (1, 2, 3, 5, 6, 7):
            sheet.cell(row=result_row, column=column).number_format = "0.000"
        for column in (4, 8):
            sheet.cell(row=result_row, column=column).number_format = "0.00%"
        sheet.cell(row=result_row, column=9).number_format = "0.00"

        if wavelength_values:
            spectrum_column = 10 + spectrum_index * 2
            spectrum_index += 1
            sheet.cell(row=2, column=spectrum_column, value=f"{current:.1f}A")
            sheet.column_dimensions[sheet.cell(row=1, column=spectrum_column).column_letter].width = 14
            sheet.column_dimensions[sheet.cell(row=1, column=spectrum_column + 1).column_letter].width = 14
            for spectrum_row, (wavelength_nm, intensity) in enumerate(
                zip(wavelength_values, intensity_values),
                start=3,
            ):
                sheet.cell(row=spectrum_row, column=spectrum_column, value=_finite_or_none(wavelength_nm))
                sheet.cell(row=spectrum_row, column=spectrum_column + 1, value=_finite_or_none(intensity))
                sheet.cell(row=spectrum_row, column=spectrum_column).number_format = "0.000000"
                sheet.cell(row=spectrum_row, column=spectrum_column + 1).number_format = "0.000000"

    temporary = target.with_name(f".{target.stem}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, target)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()


def append_test_record(path: Path, record: ExcelTestRecord) -> None:
    wavelength_values = [float(value) for value in record.wavelength]
    intensity_values = [float(value) for value in record.intensity]
    if len(wavelength_values) != len(intensity_values):
        raise ValueError("波长和强度数据长度必须一致")

    target = Path(path).expanduser()
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(target) if target.exists() else _create_workbook(target)
    sheet = workbook[workbook.sheetnames[0]]
    _ensure_result_headers(sheet)

    result_by_current: dict[float, tuple[float | None, ...]] = {}
    row = 3
    while sheet.cell(row=row, column=1).value is not None:
        current = float(sheet.cell(row=row, column=1).value)
        result_by_current[current] = tuple(sheet.cell(row=row, column=column).value for column in range(1, 10))
        row += 1
    previous_result_end_row = row - 1
    result_by_current[float(record.current_a)] = (
        record.current_a,
        record.voltage_v,
        record.power_w,
        record.efficiency,
        record.peak_wavelength_nm,
        record.centroid_nm,
        record.fwhm_nm,
        record.pib,
        record.smsr_db,
    )

    for clear_row in range(3, previous_result_end_row + 1):
        for column in range(1, 10):
            sheet.cell(row=clear_row, column=column).value = None
    for result_row, current in enumerate(sorted(result_by_current), start=3):
        for column, value in enumerate(result_by_current[current], start=1):
            sheet.cell(row=result_row, column=column, value=_finite_or_none(value) if value is not None else None)
        for column in (1, 2, 3, 5, 6, 7):
            sheet.cell(row=result_row, column=column).number_format = "0.000"
        for column in (4, 8):
            sheet.cell(row=result_row, column=column).number_format = "0.00%"
        sheet.cell(row=result_row, column=9).number_format = "0.00"

    spectra_by_current: dict[float, list[tuple[float | None, float | None]]] = {}
    previous_max_row = sheet.max_row
    previous_max_column = sheet.max_column
    for spectrum_column in range(10, previous_max_column + 1, 2):
        label = sheet.cell(row=2, column=spectrum_column).value
        if label is None:
            continue
        match = re.fullmatch(r"\s*([+-]?(?:\d+(?:\.\d*)?|\.\d+))\s*A\s*", str(label), re.IGNORECASE)
        if match is None:
            continue
        current = float(match.group(1))
        points: list[tuple[float | None, float | None]] = []
        for spectrum_row in range(3, previous_max_row + 1):
            wavelength_nm = sheet.cell(row=spectrum_row, column=spectrum_column).value
            intensity = sheet.cell(row=spectrum_row, column=spectrum_column + 1).value
            if wavelength_nm is None and intensity is None:
                continue
            points.append((wavelength_nm, intensity))
        spectra_by_current[current] = points
    if wavelength_values:
        spectra_by_current[float(record.current_a)] = list(zip(wavelength_values, intensity_values))
    else:
        spectra_by_current.pop(float(record.current_a), None)

    for clear_row in range(2, previous_max_row + 1):
        for column in range(10, previous_max_column + 1):
            sheet.cell(row=clear_row, column=column).value = None
    sheet["J1"] = "光谱"
    for spectrum_index, current in enumerate(sorted(spectra_by_current)):
        spectrum_column = 10 + spectrum_index * 2
        sheet.cell(row=2, column=spectrum_column, value=f"{current:.1f}A")
        sheet.column_dimensions[sheet.cell(row=1, column=spectrum_column).column_letter].width = 14
        sheet.column_dimensions[sheet.cell(row=1, column=spectrum_column + 1).column_letter].width = 14
        for spectrum_row, (wavelength_nm, intensity) in enumerate(spectra_by_current[current], start=3):
            sheet.cell(row=spectrum_row, column=spectrum_column, value=_finite_or_none(wavelength_nm) if wavelength_nm is not None else None)
            sheet.cell(row=spectrum_row, column=spectrum_column + 1, value=_finite_or_none(intensity) if intensity is not None else None)
            sheet.cell(row=spectrum_row, column=spectrum_column).number_format = "0.000000"
            sheet.cell(row=spectrum_row, column=spectrum_column + 1).number_format = "0.000000"

    temporary = target.with_name(f".{target.stem}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, target)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()
