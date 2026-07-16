import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from combined_test.excel_export import ExcelTestRecord, append_test_record, build_test_workbook_path, save_test_records


class ExcelExportTests(unittest.TestCase):
    def _record(self, current_a: float, test_station: str = "") -> ExcelTestRecord:
        return ExcelTestRecord(
            current_a=current_a,
            voltage_v=50.5,
            power_w=33.0,
            efficiency=33.0 / current_a / 50.5,
            peak_wavelength_nm=976.1,
            centroid_nm=976.0,
            fwhm_nm=1.2,
            pib=0.995,
            smsr_db=32.5,
            wavelength=[974.0, 976.0, 978.0],
            intensity=[10.0, 100.0, 20.0],
            test_station=test_station,
        )

    def test_path_uses_sn_station_and_minute_precision_test_time(self) -> None:
        path = build_test_workbook_path(
            Path("records"),
            "SN:001",
            datetime(2026, 7, 10, 14, 30, 25, 123456),
            "老化站:1",
        )

        self.assertEqual(path, Path("records/SN_001/老化站_1/2026_07_10_14_30.xlsx"))

    def test_filename_ignores_seconds_and_subsecond_time(self) -> None:
        first = build_test_workbook_path(
            Path("records"), "SN001", datetime(2026, 7, 10, 14, 30, 25, 1), "老化站 1"
        )
        second = build_test_workbook_path(
            Path("records"), "SN001", datetime(2026, 7, 10, 14, 30, 59, 999999), "老化站 1"
        )

        self.assertEqual(first, second)

    def test_appends_liv_and_full_spectra_to_same_reference_layout_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "SN001_2026_07_10_14_30_25.xlsx"
            append_test_record(path, self._record(4.0))
            append_test_record(path, self._record(3.0))

            workbook = load_workbook(path, data_only=False)
            sheet = workbook[workbook.sheetnames[0]]
            self.assertEqual(sheet["A1"].value, "LIV")
            self.assertEqual(sheet["J1"].value, "光谱")
            self.assertEqual([sheet.cell(2, column).value for column in range(1, 10)], [
                "电流(A)", "电压(V)", "功率(W)", "电光效率",
                "中心波长(nm)", "质心波长(nm)", "FWHM(nm)", "PIB", "SMSR(dB)",
            ])
            self.assertEqual(sheet["A3"].value, 3.0)
            self.assertEqual(sheet["A4"].value, 4.0)
            self.assertEqual(sheet["J2"].value, "3.0A")
            self.assertEqual(sheet["L2"].value, "4.0A")
            self.assertEqual([sheet.cell(row, 10).value for row in range(3, 6)], [974.0, 976.0, 978.0])
            self.assertEqual([sheet.cell(row, 11).value for row in range(3, 6)], [10.0, 100.0, 20.0])
            self.assertEqual(sheet["D3"].number_format, "0.00%")
            self.assertEqual(sheet["H3"].number_format, "0.00%")
            self.assertEqual(sheet["I3"].value, 32.5)
            self.assertEqual(sheet["I3"].number_format, "0.00")
            workbook.close()

    def test_batch_save_writes_all_records_once_in_current_order(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "batch.xlsx"

            save_test_records(path, [self._record(8.0), self._record(2.0), self._record(4.0)])

            workbook = load_workbook(path, data_only=False)
            sheet = workbook[workbook.sheetnames[0]]
            self.assertEqual([sheet.cell(row, 1).value for row in range(3, 6)], [2.0, 4.0, 8.0])
            self.assertEqual([sheet.cell(2, column).value for column in (10, 12, 14)], ["2.0A", "4.0A", "8.0A"])
            workbook.close()

    def test_batch_save_writes_test_station_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "station.xlsx"

            save_test_records(path, [self._record(2.0, "老化站 1"), self._record(4.0, "老化站 1")])

            workbook = load_workbook(path, data_only=False)
            sheet = workbook[workbook.sheetnames[0]]
            self.assertEqual(sheet["B1"].value, "测试站别")
            self.assertEqual(sheet["C1"].value, "老化站 1")
            workbook.close()

    def test_batch_save_rejects_mixed_test_stations(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "station.xlsx"

            with self.assertRaisesRegex(ValueError, "测试站别必须一致"):
                save_test_records(path, [self._record(2.0, "站别 A"), self._record(4.0, "站别 B")])

    def test_batch_save_accepts_liv_record_without_spectrum(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "without_spectrum.xlsx"
            record = ExcelTestRecord(
                current_a=3.0,
                voltage_v=50.0,
                power_w=75.0,
                efficiency=0.5,
                peak_wavelength_nm=float("nan"),
                centroid_nm=float("nan"),
                fwhm_nm=float("nan"),
                pib=float("nan"),
                smsr_db=float("nan"),
                wavelength=[],
                intensity=[],
            )

            save_test_records(path, [record])

            workbook = load_workbook(path, data_only=False)
            sheet = workbook[workbook.sheetnames[0]]
            self.assertEqual([sheet.cell(3, column).value for column in range(1, 5)], [3.0, 50.0, 75.0, 0.5])
            self.assertEqual([sheet.cell(3, column).value for column in range(5, 10)], [None] * 5)
            self.assertIsNone(sheet["J2"].value)
            workbook.close()

    def test_append_migrates_legacy_workbook_smsr_header(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "legacy.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "测试数据"
            sheet["A1"] = "LIV"
            sheet["J1"] = "光谱"
            legacy_headers = (
                "电流(A)", "电压(V)", "功率(W)", "电光效率",
                "中心波长(nm)", "质心波长(nm)", "FWHM(nm)", "PIB",
            )
            for column, header in enumerate(legacy_headers, start=1):
                sheet.cell(2, column, header)
            workbook.save(path)
            workbook.close()

            append_test_record(path, self._record(3.0))

            migrated = load_workbook(path)
            sheet = migrated["测试数据"]
            self.assertEqual(sheet["I2"].value, "SMSR(dB)")
            self.assertEqual(sheet["I3"].value, 32.5)
            migrated.close()


if __name__ == "__main__":
    unittest.main()
