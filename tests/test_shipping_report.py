import base64
import hashlib
import os
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import Workbook, load_workbook
from PySide6.QtCore import QSettings
from PySide6.QtTest import QTest
from PySide6.QtWidgets import QApplication

from combined_test.excel_export import ExcelTestRecord, save_test_records
from combined_test.record_store import SessionStatus
from combined_test.shipping_report_dialog import ShippingReportConfigurationDialog
from combined_test.shipping_report import (
    POLE_FIELD_DEFINITIONS,
    ReportFieldDefinition,
    SPECTRUM_FIELD_DEFINITIONS,
    REPORT_WORKBOOK_SCHEMA_VERSION,
    SelectedReportField,
    ShippingReportRequest,
    ShippingReportType,
    SpectrumAxisMode,
    SpectrumAxisSettings,
    _format_performance_parameters_title,
    _format_pole_summary_value,
    generate_shipping_report,
    inspect_shipping_workbook,
    load_shipping_report_preferences,
    render_shipping_report_preview,
    save_shipping_report_preferences,
    suggested_field_values,
    validate_shipping_report_request,
)


class ShippingReportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.app = QApplication.instance() or QApplication([])

    def _record(self, current_a: float, *, spectrum: bool = True) -> ExcelTestRecord:
        return ExcelTestRecord(
            current_a=current_a,
            voltage_v=10.0 + current_a,
            power_w=20.0 * current_a,
            efficiency=(20.0 * current_a) / (current_a * (10.0 + current_a)),
            peak_wavelength_nm=976.0 + current_a / 100.0 if spectrum else float("nan"),
            centroid_nm=976.0 if spectrum else float("nan"),
            fwhm_nm=0.3 if spectrum else float("nan"),
            pib=0.95 if spectrum else float("nan"),
            smsr_db=35.0 if spectrum else float("nan"),
            wavelength=[974.0, 975.0, 976.0, 977.0, 978.0] if spectrum else [],
            intensity=[10.0, 100.0, 1000.0, 120.0, 12.0] if spectrum else [],
            test_station="出货站 1",
        )

    def _session(self, status: SessionStatus = SessionStatus.COMPLETED) -> SimpleNamespace:
        return SimpleNamespace(
            session_id="session-1",
            sn="HEA000001",
            product_model="500W 976nm 光纤耦合模块",
            batch="B01",
            station="出货站 1",
            mode="automatic",
            started_at_utc="2026-07-24T01:00:00+00:00",
            ended_at_utc="2026-07-24T01:10:00+00:00",
            status=status,
            termination_reason="所有计划测试点均已保存",
            shutdown_confirmed=True,
            software_version="1.0.0",
            calculation_version="2026-07",
            settings={},
            devices=(),
        )

    def _write(self, path: Path, *, status: SessionStatus = SessionStatus.COMPLETED, mixed=False) -> None:
        records = [self._record(1.0), self._record(2.0), self._record(3.0, spectrum=not mixed)]
        save_test_records(path, records, session=self._session(status))

    def _request(self, report_type: ShippingReportType) -> ShippingReportRequest:
        definitions = SPECTRUM_FIELD_DEFINITIONS if report_type is ShippingReportType.SPECTRUM else POLE_FIELD_DEFINITIONS
        fields = {
            definition.key: SelectedReportField(
                include=definition.key in {"power", "operatingCurrent", "coolingTemperature"},
                value=("25" if definition.key == "coolingTemperature" else "3" if definition.key == "operatingCurrent" else "60"),
            )
            for definition in definitions
        }
        return ShippingReportRequest(
            report_type=report_type,
            product_name="500W 976nm 光纤耦合模块",
            sn="HEA000001",
            operating_current_a=3.0,
            fields=fields,
            legacy_completion_confirmed=False,
            spectrum_axis=(
                SpectrumAxisSettings(SpectrumAxisMode.RELATIVE_DB, -80.0, 0.0)
                if report_type is ShippingReportType.SPECTRUM
                else None
            ),
        )

    def test_approved_report_artwork_is_unchanged(self) -> None:
        assets = Path(__file__).resolve().parents[1] / "assets" / "report"
        expected_hashes = {
            "everbright_header_reference.png": "2b3f2d3cb32ce1bdf50d73f8a53ee646e3793b686cd8409a54037ac5c4343954",
            "laser_warning_reference.png": "448da5e8c41a75e166e1d5d3871886bac7494b7af8c2bb5d3f8a105c52f264f2",
        }
        for filename, expected_hash in expected_hashes.items():
            with self.subTest(filename=filename):
                self.assertEqual(hashlib.sha256((assets / filename).read_bytes()).hexdigest(), expected_hash)

    def test_inspection_accepts_verified_success_and_returns_suggestions(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "test.xlsx"
            self._write(path)

            inspection = inspect_shipping_workbook(path)

            self.assertEqual(inspection.schema_version, REPORT_WORKBOOK_SCHEMA_VERSION)
            self.assertEqual(inspection.compatibility.kind, "verified_success")
            self.assertTrue(inspection.spectrum_complete)
            self.assertEqual(inspection.allowed_report_types, ("spectrum", "pole"))
            self.assertEqual(inspection.sn, "HEA000001")
            suggestions = suggested_field_values(inspection, ShippingReportType.SPECTRUM, 3.0)
            self.assertEqual(suggestions["operatingCurrent"], "3")
            self.assertEqual(suggestions["power"], "60.0")
            self.assertEqual(suggestions["operatingVoltage"], "13.00")
            self.assertEqual(suggestions["slopeEfficiency"], "20.000")
            self.assertEqual(suggestions["electroOpticalEfficiency"], "153.85")
            pole_suggestions = suggested_field_values(inspection, ShippingReportType.POLE, 3.0)
            self.assertEqual(pole_suggestions["slopeEfficiency"], "20.000")
            zero_current = replace(inspection.point_for_current(1.0), current_a=0.0, power_w=0.0)
            self.assertIsNone(zero_current.suggestions()["slopeEfficiencyWPerA"])

    def test_pole_summary_values_use_compact_report_precision(self) -> None:
        definitions = {definition.key: definition for definition in POLE_FIELD_DEFINITIONS}

        self.assertEqual(_format_pole_summary_value(definitions["power"], "155.7"), "155.7")
        self.assertEqual(_format_pole_summary_value(definitions["operatingCurrent"], "7"), "7")
        self.assertEqual(_format_pole_summary_value(definitions["operatingVoltage"], "78.07"), "78.07")
        self.assertEqual(_format_pole_summary_value(definitions["slopeEfficiency"], "22.247"), "22.25")
        self.assertEqual(_format_pole_summary_value(definitions["electroOpticalEfficiency"], "28.50"), "28.5")
        self.assertEqual(_format_pole_summary_value(definitions["coolingTemperature"], "25"), "25")

    def test_pole_report_labels_use_the_requested_english_names(self) -> None:
        definitions = {definition.key: definition for definition in POLE_FIELD_DEFINITIONS}

        self.assertEqual(definitions["electroOpticalEfficiency"].label, "E-O efficiency")
        self.assertEqual(_format_performance_parameters_title(7.0), "Performance Parameters @7A")

    def test_explicit_failed_workbook_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "failed.xlsx"
            self._write(path, status=SessionStatus.ABORTED_SAFELY)

            inspection = inspect_shipping_workbook(path)

            self.assertEqual(inspection.compatibility.kind, "rejected")
            with self.assertRaisesRegex(ValueError, "中止或失败"):
                validate_shipping_report_request(inspection, self._request(ShippingReportType.POLE))

    def test_unknown_nonlegacy_schema_version_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "future.xlsx"
            self._write(path)
            workbook = load_workbook(path)
            sheet = workbook["会话信息"]
            for row in range(1, sheet.max_row + 1):
                if sheet.cell(row, 1).value == "出货报告格式版本":
                    sheet.cell(row, 2, "999")
                    break
            workbook.save(path)
            workbook.close()

            inspection = inspect_shipping_workbook(path)

            self.assertEqual(inspection.compatibility.kind, "rejected")
            self.assertIn("不受当前软件支持", inspection.compatibility.message)

    def test_legacy_workbook_requires_manual_confirmation(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "legacy.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "测试数据"
            sheet.cell(3, 1, 1.0)
            sheet.cell(3, 2, 12.0)
            sheet.cell(3, 3, 20.0)
            sheet.cell(3, 4, 0.5)
            workbook.save(path)
            workbook.close()
            inspection = inspect_shipping_workbook(path)
            request = self._request(ShippingReportType.POLE)

            self.assertTrue(inspection.compatibility.requires_legacy_confirmation)
            with self.assertRaisesRegex(ValueError, "确认旧版"):
                validate_shipping_report_request(inspection, request)
            confirmed = ShippingReportRequest(
                **{**request.__dict__, "operating_current_a": 1.0, "legacy_completion_confirmed": True}
            )
            validate_shipping_report_request(inspection, confirmed)

    def test_missing_one_spectrum_strictly_limits_report_type(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "mixed.xlsx"
            self._write(path, mixed=True)

            inspection = inspect_shipping_workbook(path)

            self.assertFalse(inspection.spectrum_complete)
            self.assertEqual(inspection.allowed_report_types, ("pole",))
            with self.assertRaisesRegex(ValueError, "光谱数据不完整"):
                validate_shipping_report_request(inspection, self._request(ShippingReportType.SPECTRUM))

    def test_inspection_rejects_corrupt_empty_and_duplicate_current_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            corrupt = root / "corrupt.xlsx"
            corrupt.write_bytes(b"not an xlsx workbook")
            with self.assertRaisesRegex(ValueError, "无法读取 Excel"):
                inspect_shipping_workbook(corrupt)

            empty = root / "empty.xlsx"
            Workbook().save(empty)
            with self.assertRaisesRegex(ValueError, "没有有效测试点"):
                inspect_shipping_workbook(empty)

            duplicate = root / "duplicate.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "测试数据"
            for row in (3, 4):
                sheet.cell(row, 1, 2.0)
                sheet.cell(row, 2, 12.0)
                sheet.cell(row, 3, 40.0)
                sheet.cell(row, 4, 0.5)
            workbook.save(duplicate)
            workbook.close()
            with self.assertRaisesRegex(ValueError, "重复电流点"):
                inspect_shipping_workbook(duplicate)

    def test_workpoint_switch_reprefills_measured_values_and_validation_blocks_missing_values(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsx"
            self._write(source)
            inspection = inspect_shipping_workbook(source)

            first = suggested_field_values(inspection, ShippingReportType.POLE, 1.0)
            last = suggested_field_values(inspection, ShippingReportType.POLE, 3.0)
            self.assertEqual(first["power"], "20.0")
            self.assertEqual(last["power"], "60.0")

            request = self._request(ShippingReportType.POLE)
            missing = ShippingReportRequest(
                **{
                    **request.__dict__,
                    "fields": {"power": SelectedReportField(include=True, value="")},
                }
            )
            with self.assertRaisesRegex(ValueError, "请填写已勾选参数"):
                validate_shipping_report_request(inspection, missing)

            none_selected = ShippingReportRequest(
                **{
                    **request.__dict__,
                    "fields": {"power": SelectedReportField(include=False, value="60")},
                }
            )
            with self.assertRaisesRegex(ValueError, "至少勾选一个"):
                validate_shipping_report_request(inspection, none_selected)

    def test_spectrum_axis_validation_supports_counts_and_relative_db(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsx"
            self._write(source)
            inspection = inspect_shipping_workbook(source)
            base = self._request(ShippingReportType.SPECTRUM)

            counts = ShippingReportRequest(
                **{**base.__dict__, "spectrum_axis": SpectrumAxisSettings(SpectrumAxisMode.COUNTS, 0, 1200)}
            )
            validate_shipping_report_request(inspection, counts)

            bad_counts = ShippingReportRequest(
                **{**base.__dict__, "spectrum_axis": SpectrumAxisSettings(SpectrumAxisMode.COUNTS, -1, 1200)}
            )
            with self.assertRaisesRegex(ValueError, "下限不能小于 0"):
                validate_shipping_report_request(inspection, bad_counts)

            bad_db = ShippingReportRequest(
                **{**base.__dict__, "spectrum_axis": SpectrumAxisSettings(SpectrumAxisMode.RELATIVE_DB, -50, 1)}
            )
            with self.assertRaisesRegex(ValueError, "上限不能大于 0 dB"):
                validate_shipping_report_request(inspection, bad_db)

    def test_preferences_remember_selection_and_only_manual_values(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            settings = QSettings(str(Path(temp_dir) / "report.ini"), QSettings.Format.IniFormat)
            request = self._request(ShippingReportType.SPECTRUM)

            save_shipping_report_preferences(settings, request)
            restored = load_shipping_report_preferences(settings)["spectrum"]

            self.assertIn("power", restored["selectedFields"])
            self.assertEqual(restored["manualValues"]["coolingTemperature"], "25")
            self.assertNotIn("power", restored["manualValues"])
            self.assertEqual(restored["spectrumAxis"]["minimum"], -80.0)

    def test_first_use_pole_preferences_match_the_approved_six_row_template(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            settings = QSettings(str(Path(temp_dir) / "report.ini"), QSettings.Format.IniFormat)

            preferences = load_shipping_report_preferences(settings)

            self.assertEqual(
                preferences["pole"]["selectedFields"],
                [definition.key for definition in POLE_FIELD_DEFINITIONS],
            )

    def test_custom_report_fields_are_validated_rendered_and_remembered(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsx"
            output = Path(temp_dir) / "custom.pdf"
            self._write(source)
            inspection = inspect_shipping_workbook(source)
            custom = ReportFieldDefinition("custom-package", "封装类型", "", "single")
            request = self._request(ShippingReportType.POLE)
            request = ShippingReportRequest(
                **{
                    **request.__dict__,
                    "custom_fields": (custom,),
                    "fields": {**request.fields, custom.key: SelectedReportField(True, "蝶形")},
                }
            )

            validate_shipping_report_request(inspection, request)
            generate_shipping_report(source, output, request)
            self.assertTrue(output.read_bytes().startswith(b"%PDF"))

            settings = QSettings(str(Path(temp_dir) / "report.ini"), QSettings.Format.IniFormat)
            save_shipping_report_preferences(settings, request)
            restored = load_shipping_report_preferences(settings)["pole"]
            self.assertEqual(restored["customFields"][0]["label"], "封装类型")
            self.assertIn(custom.key, restored["selectedFields"])
            self.assertEqual(restored["manualValues"][custom.key], "蝶形")

    def test_generates_both_pdf_layouts_atomically(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsx"
            self._write(source)
            for report_type in (ShippingReportType.SPECTRUM, ShippingReportType.POLE):
                output = Path(temp_dir) / f"{report_type.value}.pdf"

                result = generate_shipping_report(source, output, self._request(report_type))

                self.assertEqual(result, output.resolve())
                self.assertTrue(output.read_bytes().startswith(b"%PDF"))
                self.assertGreater(output.stat().st_size, 10_000)
                self.assertFalse((Path(temp_dir) / f".{report_type.value}.tmp.pdf").exists())

    def test_renders_live_preview_pages_from_the_export_layout(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsx"
            self._write(source)

            pages = render_shipping_report_preview(source, self._request(ShippingReportType.POLE))

            self.assertEqual(len(pages), 1)
            self.assertTrue(base64.b64decode(pages[0]).startswith(b"\x89PNG\r\n\x1a\n"))

    def test_qt_configuration_dialog_refreshes_the_preview_after_an_edit(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.xlsx"
            self._write(source)
            dialog = ShippingReportConfigurationDialog(
                inspect_shipping_workbook(source),
                QSettings(str(Path(temp_dir) / "report.ini"), QSettings.Format.IniFormat),
            )
            dialog.show()
            self.assertIn("slopeEfficiency", dialog.field_widgets)
            self.assertEqual(dialog.field_widgets["slopeEfficiency"][0].text(), "Slope Efficiency")
            dialog.field_widgets["power"][0].setChecked(True)
            QTest.qWait(1_200)

            self.assertTrue(dialog._preview_pixmaps)
            self.assertIn("与导出的 PDF 一致", dialog.preview_status_label.text())
            dialog.close()

    def test_schema_version_is_written_to_session_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "schema.xlsx"
            self._write(path)
            workbook = load_workbook(path, data_only=True)
            values = {
                str(workbook["会话信息"].cell(row, 1).value): str(workbook["会话信息"].cell(row, 2).value)
                for row in range(1, workbook["会话信息"].max_row + 1)
            }
            workbook.close()
            self.assertEqual(values["出货报告格式版本"], REPORT_WORKBOOK_SCHEMA_VERSION)


if __name__ == "__main__":
    unittest.main()
